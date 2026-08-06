#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""TSV から配布用のエクセル（excel/ 配下）を生成する。

使い方:
    python3 build_excel.py

TSVを変更したら実行し直すこと。エクセルを直接編集してもTSVには反映されない。
必要: openpyxl（pip3 install openpyxl）
"""
import io, os, collections
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter

B = os.path.dirname(os.path.abspath(__file__))
OUT = os.path.join(B, 'excel')
os.makedirs(OUT, exist_ok=True)

HEAD_FILL = PatternFill('solid', fgColor='2F4858')
HEAD_FONT = Font(color='FFFFFF', bold=True, size=11)
THIN = Side(style='thin', color='CCCCCC')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
FLOW_COLOR = {'1': 'FFF2CC', '2': 'DEEAF6', '3': 'E2EFDA', '4': 'FCE4D6'}

def load(name):
    p = os.path.join(B, name)
    return [r.split('\t') for r in io.open(p, encoding='utf-8').read().rstrip('\n').split('\n')]

def build_sheet(ws, rows, widths):
    for r in rows:
        ws.append(r)
    for i, w in enumerate(widths[:len(rows[0])], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for c in ws[1]:
        c.fill = HEAD_FILL
        c.font = HEAD_FONT
        c.alignment = Alignment(vertical='center', wrap_text=True)
    ws.row_dimensions[1].height = 30
    for row in ws.iter_rows(min_row=2):
        for c in row:
            c.alignment = Alignment(vertical='top', wrap_text=True)
            c.border = BORDER
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions

    H = rows[0]
    col = lambda nm: H.index(nm) + 1 if nm in H else None
    for nm, opts in [('実装状況', '"○,△,×"'), ('判定', '"合格,不合格"')]:
        i = col(nm)
        if i:
            dv = DataValidation(type='list', formula1=opts, allow_blank=True)
            ws.add_data_validation(dv)
            dv.add(f'{get_column_letter(i)}2:{get_column_letter(i)}{len(rows)}')
    i = col('フロー')
    if i:
        rng = f'{get_column_letter(i)}2:{get_column_letter(i)}{len(rows)}'
        for k, color in FLOW_COLOR.items():
            ws.conditional_formatting.add(rng, CellIsRule(
                operator='equal', formula=[f'"{k}"'], fill=PatternFill('solid', fgColor=color)))
    i = col('MVP区分')
    if i:
        rng = f'{get_column_letter(i)}2:{get_column_letter(i)}{len(rows)}'
        ws.conditional_formatting.add(rng, CellIsRule(
            operator='equal', formula=['"MVP必須"'],
            fill=PatternFill('solid', fgColor='FFE0E0'), font=Font(bold=True)))
    j = col('判定')
    if j:
        rng = f'{get_column_letter(j)}2:{get_column_letter(j)}{len(rows)}'
        ws.conditional_formatting.add(rng, CellIsRule(operator='equal', formula=['"合格"'],
            fill=PatternFill('solid', fgColor='D6F5D6')))
        ws.conditional_formatting.add(rng, CellIsRule(operator='equal', formula=['"不合格"'],
            fill=PatternFill('solid', fgColor='FFD6D6')))

FILES = [
    ('01_受け入れ条件_使い方.xlsx', '使い方', '20260804_受け入れ条件_使い方.tsv', [10, 22, 60, 50]),
    ('02_受け入れ条件_定義表.xlsx', '定義表', '20260804_受け入れ条件_定義表.tsv', [18, 10, 8, 10, 34, 12, 52, 30]),
    ('03_受け入れ条件_確認表.xlsx', '確認表', '20260804_受け入れ条件_確認表.tsv', [10, 8, 12, 10, 30, 20, 52, 22, 18, 10, 34, 8, 10, 20]),
    ('04_検証プラン_操作系.xlsx', '操作系', '20260803_検証プラン_操作系.tsv', [14, 10, 8, 12, 24, 52, 52, 14, 44, 28, 8, 46]),
    ('05_検証プラン_計算系.xlsx', '計算系', '20260803_検証プラン_計算系.tsv', [10, 8, 10, 24, 34, 44, 14, 44, 28, 8, 46]),
    ('06_検証用データセット.xlsx', 'データセット', '20260803_検証用データセット.tsv', [12, 8, 12, 8, 60, 40]),
    ('07_実装マイルストーン表.xlsx', 'マイルストーン', '20260804_実装マイルストーン表.tsv', [10, 24, 40, 10, 12, 8, 40, 50, 10, 10]),
    ('08_トレーサビリティ表.xlsx', 'トレーサビリティ', '20260803_トレーサビリティ表.tsv', [26, 44, 14, 20, 8]),
]

for out, name, src, widths in FILES:
    wb = Workbook()
    ws = wb.active
    ws.title = name
    rows = [r for r in load(src) if r[0] != '#']
    build_sheet(ws, rows, widths)
    wb.save(os.path.join(OUT, out))
    print(f'  {out:34} {len(rows)-1:>3}行 × {len(rows[0]):>2}列')

# 09: 進捗（集計は同一ファイル内のシートしか参照できないため確認表を同梱）
wb = Workbook()
ws = wb.active
ws.title = '進捗'
for c, w in zip('ABCD', [30, 14, 14, 42]):
    ws.column_dimensions[c].width = w
ws.append(['タイムレコード 登録・編集機能　進捗'])
ws['A1'].font = Font(bold=True, size=14)
ws.append([])
ws.append(['フロー', 'MVP必須 合格', 'MVP必須 総数', '内容'])
for c in ws[3]:
    c.fill = HEAD_FILL
    c.font = HEAD_FONT
FLOWS = {'1': '新規登録（手入力モーダル）', '2': '新規登録（CSVインポート）',
         '3': '運用の登録・編集（手入力）', '4': '運用の登録（CSV）'}
for f, nm in FLOWS.items():
    ws.append([f'フロー{f}',
        f'=COUNTIFS(確認表!B:B,"*{f}*",確認表!C:C,"MVP必須",確認表!L:L,"合格")',
        f'=COUNTIFS(確認表!B:B,"*{f}*",確認表!C:C,"MVP必須")', nm])
ws.append([])
ws.append(['マイルストーン', 'MVP必須 合格', 'MVP必須 総数', '完了とみなす条件'])
for c in ws[9]:
    c.fill = HEAD_FILL
    c.font = HEAD_FONT
MS = [('M1', '配布物とアップロードの入口／手入力の入口'), ('M2', 'プレビューと判定／登録モーダル'),
      ('M3', '実行と完了'), ('M4', '書き込みの正しさ'),
      ('M5', '運用更新と再取り込み'), ('M6', '権限と手入力との分担')]
for m, nm in MS:
    ws.append([f'{m} {nm}',
        f'=COUNTIFS(確認表!D:D,"{m}",確認表!C:C,"MVP必須",確認表!L:L,"合格")',
        f'=COUNTIFS(確認表!D:D,"{m}",確認表!C:C,"MVP必須")', ''])
ws.append([])
ws.append(['MVP必須 合計', '=COUNTIFS(確認表!C:C,"MVP必須",確認表!L:L,"合格")',
           '=COUNTIF(確認表!C:C,"MVP必須")', 'すべて合格でリリース可'])
ws.append(['リリース必須 合計', '=COUNTIFS(確認表!C:C,"リリース必須",確認表!L:L,"合格")',
           '=COUNTIF(確認表!C:C,"リリース必須")', '不合格は影響を確認して判断'])
ws.append(['不合格（全体）', '=COUNTIF(確認表!L:L,"不合格")', '', 'MVP必須に1件でもあればリリース不可'])
for r in ws.iter_rows(min_row=3):
    for c in r:
        c.border = BORDER
ws2 = wb.create_sheet('確認表')
build_sheet(ws2, [r for r in load('20260804_受け入れ条件_確認表.tsv') if r[0] != '#'],
            [10, 8, 12, 10, 30, 20, 52, 22, 18, 10, 34, 8, 10, 20])
wb.save(os.path.join(OUT, '09_進捗（確認表つき）.xlsx'))
print('  09_進捗（確認表つき）.xlsx           集計 + 確認表')
print(f'\n出力先: {OUT}')
