# -*- coding: utf-8 -*-
"""TSV → 配布用エクセル（01・02）を同期する"""
import io,os,re
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font,Alignment,PatternFill,Border,Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.cell.rich_text import CellRichText,TextBlock
from openpyxl.cell.text import InlineFont
from openpyxl.utils import get_column_letter
os.chdir(os.path.expanduser('~/Projects/rakmy-timerecord-import-spec'))
D=os.path.expanduser('~/Desktop/タイムレコード_20260805')
def rich(s):
    if not isinstance(s,str) or '**' not in s: return s
    parts=re.split(r'\*\*(.+?)\*\*', s); out=[]
    for i,x in enumerate(parts):
        if x: out.append(TextBlock(InlineFont(b=True,sz=10) if i%2 else InlineFont(sz=10), x))
    return CellRichText(*out)
def load(p):
    r=[x.split('\t') for x in io.open(p,encoding='utf-8').read().rstrip('\n').split('\n')]
    return {n:i for i,n in enumerate(r[0])},[x for x in r[1:] if x[0]!='#']

# ---- 01 受入条件表（書式・集計式を保って上書き）----
hc,C=load('20260804_受け入れ条件_確認表.tsv'); hd,Dd=load('20260804_受け入れ条件_定義表.tsv')
c={r[0]:r for r in C}; d={r[hd['条件ID']]:r for r in Dd}
p1=os.path.join(D,'20260805_タイムレコード_01_受入条件表.xlsx')
wb=load_workbook(p1, rich_text=True); ws=wb.active
hx={str(ws.cell(28,i).value):i for i in range(1,ws.max_column+1)}
SRC={'設計ポイント':lambda k:d[k][hd['設計ポイント']],'実装レベル':lambda k:c[k][hc['MVP区分']],
 '何を実装するか':lambda k:c[k][hc['満たすべきこと']],
 '合格ライン（この数値を満たせば実装完了）':lambda k:c[k][hc['合致していると言える状態']],
 '実装できていないと起きること':lambda k:d[k][hd['崩れると起きること']],
 'フロー':lambda k:c[k][hc['フロー']],'確認する画面':lambda k:c[k][hc['実装箇所（画面・機能）']],
 '確認方法':lambda k:c[k][hc['確認方法']],'仕様書の根拠':lambda k:d[k][hd['仕様の該当箇所']],
 '対応検証ID':lambda k:c[k][hc['根拠（検証ID）']],'ヘルプ該当箇所':lambda k:c[k][hc['ヘルプ該当箇所']]}
n1=0
for r in range(29,ws.max_row+1):
    k=ws.cell(r,hx['条件ID']).value
    if k not in c: continue
    for col,fn in SRC.items():
        if col not in hx: continue
        want=fn(k)
        if str(ws.cell(r,hx[col]).value or '')!=want.replace('**',''):
            ws.cell(r,hx[col]).value=rich(want); n1+=1
wb.save(p1)

# ---- 02 検証プラン（作り直し）----
o,OP=load('20260803_検証プラン_操作系.tsv'); a,CA=load('20260803_検証プラン_計算系.tsv')
COLS=[('段階',15),('検証ID',10),('フロー',8),('検証内容',26),('先に実施する検証',24),
      ('初期設定（前提データ・データセット）',36),('操作・前提条件／入力値',46),('期待挙動（判定基準）',50),
      ('スクショファイル名',30),('証跡（撮る画面と状態）',42),('画面（どこを開いて操作するか）',46),
      ('確認ポイント',34),('対応する条件ID',16),('ヘルプ該当箇所',42),('実際の値',24),('判定',10),('計算根拠・備考',24)]
def row(ix,r):
    ope='操作・前提条件' if '操作・前提条件' in ix else '入力値'
    v=r[ix[ope]] if ope=='操作・前提条件' else '入力値: '+r[ix[ope]]
    last='根拠・備考' if '根拠・備考' in ix else '計算根拠'
    return [r[ix['段階']],r[ix['検証ID']],r[ix['フロー']],r[ix['検証内容']],r[ix['先に実施する検証']],
            r[ix['初期設定（前提データ）']],v,r[ix['期待挙動（判定基準）']],
            r[ix['スクショファイル名']],r[ix['証跡（撮る画面と状態）']],r[ix['画面（どこを開いて操作するか）']],
            r[ix['確認ポイント']],r[ix['対応する条件ID']],r[ix['ヘルプ該当箇所']],'','',r[ix[last]]]
rows=[row(o,r) for r in OP]+[row(a,r) for r in CA]
wb=Workbook(); ws=wb.active; ws.title='検証プラン'
HF=PatternFill('solid',fgColor='2F4858'); TH=Side(style='thin',color='CCCCCC')
BD=Border(left=TH,right=TH,top=TH,bottom=TH); YEL=PatternFill('solid',fgColor='FFF9C4')
KEY=PatternFill('solid',fgColor='EAF1F6'); LNK=PatternFill('solid',fgColor='F2F0E4'); SHT=PatternFill('solid',fgColor='FDEBD3')
ws.cell(1,1,'確かめる手順（85ケース）。「初期設定」のデータを用意し、「画面」を開いて「操作」を行い、「期待挙動」と一致するかを見る。'
            '期待挙動の【　】は判定する画面。「証跡」の状態を撮って「スクショファイル名」で保存する。'
            '結果は「対応する条件ID」で判定し、仕様が変われば「ヘルプ該当箇所」を直す。').font=Font(size=10)
for i,(nm,w) in enumerate(COLS,1):
    cc=ws.cell(2,i,nm); cc.fill=HF; cc.font=Font(color='FFFFFF',bold=True,size=10)
    cc.alignment=Alignment(vertical='center',wrap_text=True); ws.column_dimensions[get_column_letter(i)].width=w
ws.row_dimensions[2].height=36
for ri,r in enumerate(rows,3):
    for ci,v in enumerate(r,1):
        cc=ws.cell(ri,ci,rich(v)); cc.alignment=Alignment(vertical='top',wrap_text=True); cc.border=BD
    for ci in (6,7,8,11): ws.cell(ri,ci).fill=KEY
    for ci in (9,10): ws.cell(ri,ci).fill=SHT
    for ci in (13,14): ws.cell(ri,ci).fill=LNK
    for ci in (15,16,17): ws.cell(ri,ci).fill=YEL
nn=len(rows)+2
dv=DataValidation(type='list',formula1='"OK,NG,対象外"',allow_blank=True); ws.add_data_validation(dv); dv.add(f'P3:P{nn}')
ws.freeze_panes='D3'; ws.auto_filter.ref=f'A2:Q{nn}'
wb.save(os.path.join(D,'20260805_タイムレコード_02_検証プラン.xlsx'))
import shutil
for f in ['20260805_タイムレコード_01_受入条件表.xlsx','20260805_タイムレコード_02_検証プラン.xlsx',
          '20260805_タイムレコード_03_検証用データセット.xlsx','20260805_タイムレコード_00_定義表.xlsx',
          '20260805_タイムレコード_使い方.xlsx']:
    shutil.copy(os.path.join(D,f),'excel/')
print(f'01: {n1}セル同期 ／ 02: {len(rows)}件 再生成 ／ excel/ へコピー')

# ---- ヘルプページ.xlsx（掲載本文＋フロー図＋掲載前チェック）----
def build_help():
    from openpyxl.drawing.image import Image as XLImage
    md = io.open('20260804_Notion掲載用_CSVで一括登録する.md', encoding='utf-8').read()
    src = io.open('20260803_ヘルプページ_CSV一括登録_ドラフト.md', encoding='utf-8').read()
    HF = PatternFill('solid', fgColor='2F4858'); HFo = Font(color='FFFFFF', bold=True, size=10)
    H1F = PatternFill('solid', fgColor='2F4858'); H2F = PatternFill('solid', fgColor='DCE6F1')
    H3F = PatternFill('solid', fgColor='EDF2F8')
    def clean(s):
        return re.sub(r'`(.+?)`', r'\1', re.sub(r'\*\*(.+?)\*\*', r'\1', s)).strip()
    L = md.split('\n'); rows = []; para = []; i = 0; shot = 0
    def flush():
        nonlocal para
        if para:
            rows.append(['', '', clean('\n'.join(para))]); para = []
    while i < len(L):
        l = L[i]
        if l.startswith('# '):
            flush(); rows.append(['H1', clean(l[2:]), '']); i += 1; continue
        if l.startswith('## '):
            flush(); rows.append(['H2', clean(l[3:]), '']); i += 1; continue
        if l.startswith('### '):
            flush(); rows.append(['H3', clean(l[4:]), '']); i += 1; continue
        if l.startswith('> 🖼'):
            flush(); rows.append(['フロー図', '', clean(l[2:]).replace('🖼', '').strip()]); i += 1; continue
        if l.startswith('> 📸'):
            flush(); shot += 1
            rows.append(['画像', f'S-{shot:02d}',
                         clean(l[2:]).replace('📸', '').replace('スクリーンショット挿入位置：', '').strip()])
            i += 1; continue
        if l.startswith('>'):
            flush(); rows.append(['注意', '', clean(l.lstrip('> '))]); i += 1; continue
        if l.startswith('|') and i + 1 < len(L) and re.match(r'^\|[\s:|-]+\|$', L[i + 1].strip()):
            flush(); head = [clean(x) for x in l.strip('|').split('|')]; i += 2
            rows.append(['表', ' / '.join(x for x in head if x), ''])
            while i < len(L) and L[i].startswith('|'):
                cells = [clean(x) for x in L[i].strip('|').split('|')]
                rows.append(['表の行', '', '　│　'.join(f'{a}：{b}' for a, b in zip(head, cells) if b)])
                i += 1
            continue
        if l.strip().startswith('```'):
            flush(); i += 1; code = []
            while i < len(L) and not L[i].strip().startswith('```'):
                code.append(L[i]); i += 1
            i += 1; rows.append(['図', '', '\n'.join(code)]); continue
        if l.strip() in ('---', ''):
            flush(); i += 1; continue
        para.append(l); i += 1
    flush()
    wb = Workbook(); ws = wb.active; ws.title = 'マニュアル本文'
    for i2, w in enumerate([10, 38, 88, 8], 1):
        ws.column_dimensions[get_column_letter(i2)].width = w
    for i2, v in enumerate(['レベル', '見出し／画像番号', '本文', '掲載'], 1):
        c = ws.cell(1, i2, v); c.fill = HF; c.font = HFo
        c.alignment = Alignment(vertical='center', wrap_text=True)
    YEL2 = PatternFill('solid', fgColor='FFF9C4')
    for r, v in enumerate(rows, 2):
        for j, x in enumerate(v + [''], 1):
            c = ws.cell(r, j, rich(x) if isinstance(x, str) else x)
            c.alignment = Alignment(vertical='top', wrap_text=True); c.border = BD
        if v[0] == 'H1':
            for j in range(1, 5):
                ws.cell(r, j).fill = H1F; ws.cell(r, j).font = Font(bold=True, color='FFFFFF', size=12)
        elif v[0] == 'H2':
            for j in range(1, 5):
                ws.cell(r, j).fill = H2F; ws.cell(r, j).font = Font(bold=True, size=11)
        elif v[0] == 'H3':
            for j in range(1, 5):
                ws.cell(r, j).fill = H3F; ws.cell(r, j).font = Font(bold=True)
        elif v[0] == '注意':
            ws.cell(r, 1).fill = PatternFill('solid', fgColor='FFF9C4')
        elif v[0] == '画像':
            ws.cell(r, 1).fill = PatternFill('solid', fgColor='FFE0B2')
        elif v[0] == 'フロー図':
            ws.cell(r, 1).fill = PatternFill('solid', fgColor='C8E6C9')
        ws.cell(r, 4).fill = YEL2
    dv2 = DataValidation(type='list', formula1='"✓"', allow_blank=True)
    ws.add_data_validation(dv2); dv2.add(f'D2:D{len(rows) + 1}')
    ws.freeze_panes = 'C2'; ws.auto_filter.ref = f'A1:D{len(rows) + 1}'
    wsf = wb.create_sheet('フロー図'); wsf.column_dimensions['A'].width = 110
    wsf['A1'] = '本文に挿入する図です。PNG と SVG がヘルプページのフォルダにあります。'
    wsf['A1'].font = Font(size=12, bold=True)
    pos = 3
    for name, ttl in [('フロー図_登録の順番.png', '1番　はじめて登録するときの順番'),
                      ('フロー図_CSVの手順.png', '6番　CSVで取り込むときの5つの手順')]:
        fp = os.path.join(D, 'ヘルプページ', name)
        wsf[f'A{pos}'] = ttl; wsf[f'A{pos}'].font = Font(size=13, bold=True, color='2F4858')
        if os.path.exists(fp):
            img = XLImage(fp); img.width, img.height = 700, 700
            wsf.add_image(img, f'A{pos + 1}')
        pos += 40
    ws3 = wb.create_sheet('公開手順とチェック')
    ws3.column_dimensions['A'].width = 92; ws3.column_dimensions['B'].width = 8
    ws3['A1'] = '本ページは、CSVインポートと初期セットアップが実装・検証を通ってから公開します。'
    ws3['A1'].font = Font(size=11, bold=True, color='B00000')
    for i2, v in enumerate(['確認項目', '完了'], 1):
        c = ws3.cell(3, i2, v); c.fill = HF; c.font = HFo
    cr = 4
    for x in re.findall(r'^- \[ \] (.+)$', src, re.M):
        ws3.cell(cr, 1, rich(x)).alignment = Alignment(vertical='top', wrap_text=True)
        ws3.cell(cr, 1).border = BD; ws3.cell(cr, 2).fill = YEL2; ws3.cell(cr, 2).border = BD
        cr += 1
    dv3 = DataValidation(type='list', formula1='"✓"', allow_blank=True)
    ws3.add_data_validation(dv3); dv3.add(f'B4:B{cr - 1}')
    ws3.freeze_panes = 'A4'
    wb.save(os.path.join(D, 'ヘルプページ', 'ヘルプページ.xlsx'))
    shutil.copy(os.path.join(D, 'ヘルプページ', 'ヘルプページ.xlsx'), 'excel/')
    return len(rows), cr - 4


_hn, _hc2 = build_help()
print(f'ヘルプページ.xlsx: 本文{_hn}行 / 掲載前チェック{_hc2}件')
