#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""全資料の整合を機械的に検査する。

使い方:
    python3 check_all.py

TSV・仕様書・ヘルプ・配布用エクセル・デスクトップの配布物を横断して、
数字の食い違いや参照切れを検出する。終了コード0で整合。

**資料を1つでも直したら、必ずこれを実行すること。**
件数や列数は複数の資料に複製されているため、1箇所直すと他が古くなる。
"""
import io, os, re, sys, collections

B = os.path.dirname(os.path.abspath(__file__))
DESK = os.path.expanduser('~/Desktop/タイムレコード_20260805')
NG = []


def ng(category, msg):
    NG.append((category, msg))


def load(name):
    p = os.path.join(B, name)
    rows = [r.split('\t') for r in io.open(p, encoding='utf-8').read().rstrip('\n').split('\n')]
    return rows[0], [r for r in rows[1:] if r[0] != '#']


def read(name, base=B):
    p = os.path.join(base, name)
    return io.open(p, encoding='utf-8').read() if os.path.exists(p) else ''


# ---------------------------------------------------------------- 1. 受け入れ条件
hd, D = load('20260804_受け入れ条件_定義表.tsv')
hc, C = load('20260804_受け入れ条件_確認表.tsv')
d = {r[1]: r for r in D}
c = {r[0]: r for r in C}

if set(d) != set(c):
    ng('受け入れ条件', f'定義表と確認表で条件IDが違う: {sorted(set(d) ^ set(c))}')
for k in sorted(set(d) & set(c)):
    for name, x, y in [('フロー', d[k][2], c[k][1]), ('MVP区分', d[k][5], c[k][2]),
                       ('マイルストーン', d[k][3], c[k][3]), ('満たすべきこと', d[k][4], c[k][4])]:
        if x.strip() != y.strip():
            ng('受け入れ条件', f'{k} の{name}が2表で違う（定義表「{x}」／確認表「{y}」）')

for label, rows, i in [('満たすべきこと', D, 4), ('合致していると言える状態', C, 6)]:
    g = collections.defaultdict(list)
    for r in rows:
        g[r[i].strip()].append(r[1] if rows is D else r[0])
    for text, ids in g.items():
        if len(ids) > 1:
            ng('受け入れ条件', f'{label}が同じ条件が複数: {" ".join(ids)} → {text[:40]}')

# 条件の本文が参照する条件IDが実在するか
for r in C:
    for m in re.findall(r'AC-\d+', r[6]):
        if m not in c:
            ng('受け入れ条件', f'{r[0]} の合致条件が存在しない {m} を参照している')

# ---------------------------------------------------------------- 2. 検証プラン
ho, OP = load('20260803_検証プラン_操作系.tsv')
ha, CA = load('20260803_検証プラン_計算系.tsv')
hds, DS = load('20260803_検証用データセット.tsv')


def col(header, name):
    """列名から位置を引く。見つからなければ検査自体を異常にする（空振り防止）"""
    if name not in header:
        ng('検査の不備', f'列「{name}」が見つからない。列名を変えたら check_all.py も直すこと')
        return None
    return header.index(name)


io_, ic_ = col(ho, '検証ID'), col(ha, '検証ID')
ids_ = col(hds, 'データセットID')
iuse = col(hds, '用途・対応する検証ID')
ievi = col(hc, '根拠（検証ID）')
vids = ({r[io_] for r in OP} if io_ is not None else set()) | ({r[ic_] for r in CA} if ic_ is not None else set())
dids = {r[ids_] for r in DS} if ids_ is not None else set()

used = collections.Counter()
for r in C:
    for m in re.findall(r'(?:OP|CA)-\d+', r[ievi]):
        used[m] += 1
        if m not in vids:
            ng('検証プラン', f'{r[0]} の根拠にある {m} が検証プランに無い')
for v in sorted(vids - set(used)):
    ng('検証プラン', f'{v} はどの受け入れ条件からも参照されていない')

for r in OP + CA:
    for m in re.findall(r'D\d+-\d+', ' '.join(r)):
        if m not in dids:
            ng('検証データ', f'検証プランが存在しない {m} を参照している')
for r in DS:
    for m in re.findall(r'(?:OP|CA)-\d+', r[iuse]):
        if m not in vids:
            ng('検証データ', f'{r[0]} の用途にある {m} が検証プランに無い')

# ---------------------------------------------------------------- 3. 実データの数
lv = collections.Counter(r[2] for r in C)
N = {
    '受け入れ条件': len(C),
    'MVP必須': lv['MVP必須'],
    'リリース必須': lv['リリース必須'],
    '改善': lv['改善'],
    '操作系': len(OP),
    '計算系': len(CA),
    '検証項目': len(OP) + len(CA),
    'データセット': len(DS),
    '操作系の列': len(ho),
    '計算系の列': len(ha),
}

# ---------------------------------------------------------------- 4. 資料に書かれた数と突き合わせ
spec = read('20260803_CSV一括登録機能_仕様書（PdM向け）.md')
CHECKS = [
    ('仕様書', r'\*\*何を守るか・なぜか\*\*（(\d+)件）', N['受け入れ条件']),
    ('仕様書', r'\*\*実装が条件に合致しているかの確認\*\*（(\d+)件）', N['受け入れ条件']),
    ('仕様書', r'エンジニアはMVP必須の(\d+)件を先に潰して', N['MVP必須']),
    ('仕様書', r'\| MVP必須(\d+)件がすべて合格 \|', N['MVP必須']),
    ('仕様書', r'操作手順と期待挙動(\d+)件', N['操作系']),
    ('仕様書', r'数値・件数・上限の確認(\d+)件', N['計算系']),
    ('仕様書', r'検証に使うデータ(\d+)件', N['データセット']),
    ('仕様書', r'検証項目は(\d+)件です', N['検証項目']),
    ('仕様書', r'操作系（(\d+)列）', N['操作系の列']),
    ('仕様書', r'計算系（(\d+)列）', N['計算系の列']),
]
for doc, pat, want in CHECKS:
    got_all = re.findall(pat, spec)
    if not got_all:
        ng('検査の不備', f'{doc}で「{pat}」が1件も一致しない。文言を変えたら check_all.py も直すこと')
    for got in got_all:
        if int(got) != want:
            ng(doc, f'「{pat}」が {got} だが実データは {want}')

# 実装レベルの件数表
m = re.search(r'\| \*\*MVP必須\*\* \|[^|]+\| \*\*(\d+)件\*\* \|', spec)
if not m:
    ng('検査の不備', '仕様書のMVP区分の表が見つからない（検査が空振りしている）')
elif int(m.group(1)) != N['MVP必須']:
    ng('仕様書', f'MVP区分の表が {m.group(1)}件 だが実データは {N["MVP必須"]}件')

# フロー別の件数
for f in '1234':
    rows = [r for r in C if f in [x.strip() for x in r[1].split('/')]]
    mvp = sum(1 for r in rows if r[2] == 'MVP必須')
    pat = r'\| \*\*' + f + r'\*\* \| [^|]+ \| (\d+)件 \| \*\*(\d+)件\*\* \| (\d+)件 \|'
    mm = re.search(pat, spec)
    if not mm:
        ng('検査の不備', f'仕様書のフロー{f}の件数行が見つからない（検査が空振りしている）')
    if mm:
        if int(mm.group(1)) != len(rows):
            ng('仕様書', f'フロー{f}の受け入れ条件が {mm.group(1)}件 だが実データは {len(rows)}件')
        if int(mm.group(2)) != mvp:
            ng('仕様書', f'フロー{f}のMVP必須が {mm.group(2)}件 だが実データは {mvp}件')

# ---------------------------------------------------------------- 5. ヘルプ（顧客に出す本文）
help_md = read('20260804_Notion掲載用_CSVで一括登録する.md')
draft = read('20260803_ヘルプページ_CSV一括登録_ドラフト.md')
for word, why in [('⚠️要確認', '実装未確定の内部メモ'), ('要確認：', '内部メモ'),
                  ('TODO', '作業メモ'), ('掲載しません', '作業用の節')]:
    if word in help_md:
        ng('ヘルプ', f'掲載本文に「{word}」が残っている（{why}）')
for ch in '①②③④⑤⑥⑦⑧⑨':
    if ch in help_md:
        ng('ヘルプ', f'掲載本文に丸数字「{ch}」が残っている（表示環境で潰れる）')
        break
if '業態・ブランド' in help_md:
    ng('ヘルプ', '掲載本文に旧表記「業態・ブランド」がある（2026-06-24 に業態へ統合）')
# 節番号の連番
nums = [int(m) for m in re.findall(r'^## (\d+)\. ', help_md, re.M)]
if nums != list(range(1, len(nums) + 1)):
    ng('ヘルプ', f'節番号が連番でない: {nums}')
# 本文が参照する節番号が実在するか
for m in set(re.findall(r'「(\d+)\. ', help_md)):
    if int(m) not in nums:
        ng('ヘルプ', f'本文が存在しない節「{m}番」を参照している')

# ---------------------------------------------------------------- 6. 配布用エクセルの数字
def xlsx_cells(name):
    """配布用エクセルの全セルを (シート名, 行, 列, 値) で返す。openpyxl が無ければ空。"""
    p = os.path.join(DESK, name)
    if not os.path.exists(p):
        return None
    try:
        from openpyxl import load_workbook
    except ImportError:
        return None
    wb = load_workbook(p)
    out = []
    for ws in wb:
        for row in ws.iter_rows():
            for c in row:
                if c.value is not None:
                    out.append((ws.title, c.row, c.column, c.value))
    return out


def xlsx_pair(name, label, want, col=2):
    """A列が label の行の col 列が want と一致するか。"""
    cells = xlsx_cells(name)
    if cells is None:
        return
    byrow = {}
    for _, r, cc, v in cells:
        byrow.setdefault(r, {})[cc] = v
    found = False
    for r, d in byrow.items():
        if str(d.get(1, '')).strip() == label:
            found = True
            got = d.get(col)
            if str(got).strip() != str(want):
                ng('配布エクセル', f'{name} の「{label}」が {got} だが実データは {want}')
    if not found:
        ng('配布エクセル', f'{name} に「{label}」の行が見つからない（検査が空振りしている）')


ms = collections.Counter(r[3] for r in C)
if os.path.isdir(DESK):
    for label, want in [('MVP必須', N['MVP必須']), ('リリース必須', N['リリース必須']),
                        ('改善', N['改善'])]:
        xlsx_pair('20260805_タイムレコード_00_定義表.xlsx', label, want)
    for f in '1234':
        rows = sum(1 for r in C if f in [x.strip() for x in r[1].split('/')])
        cells = xlsx_cells('20260805_タイムレコード_00_定義表.xlsx') or []
        hit = [v for _, r, cc, v in cells if cc == 1 and isinstance(v, str) and v.startswith(f'フロー{f} ')]
        if not hit:
            ng('配布エクセル', f'00_定義表.xlsx に「フロー{f}」の行が無い（検査が空振りしている）')
        else:
            byrow = {}
            for _, r, cc, v in cells:
                byrow.setdefault(r, {})[cc] = v
            for r, d in byrow.items():
                if isinstance(d.get(1), str) and d[1].startswith(f'フロー{f} '):
                    if str(d.get(2)).strip() != str(rows):
                        ng('配布エクセル', f'00_定義表.xlsx のフロー{f}が {d.get(2)} だが実データは {rows}')
    # 使い方
    use = xlsx_cells('20260805_タイムレコード_使い方.xlsx')
    if use is not None:
        text = '\n'.join(str(v) for _, _, _, v in use)
        m = re.search(r'MVP必須(\d+)件をエンジニアと合意', text)
        if not m:
            ng('配布エクセル', '使い方.xlsx に「MVP必須N件をエンジニアと合意」が無い（検査が空振りしている）')
        elif int(m.group(1)) != N['MVP必須']:
            ng('配布エクセル', f'使い方.xlsx のMVP必須が {m.group(1)}件 だが実データは {N["MVP必須"]}件')
        m = re.search(r'(\d+)〜(\d+)行 ?│? ?受入条件(\d+)件', text.replace(' ', ''))
        if f'受入条件{N["受け入れ条件"]}件' not in text.replace(' ', ''):
            ng('配布エクセル', f'使い方.xlsx の受入条件件数が {N["受け入れ条件"]}件 と書かれていない')

# ---------------------------------------------------------------- 7. 配布物の同期
PAIRS = [
    ('ヘルプページ/ヘルプページ_掲載用.md', '20260804_Notion掲載用_CSVで一括登録する.md'),
    ('Notion貼り付け用_md/02_ヘルプ_本文.md', '20260804_Notion掲載用_CSVで一括登録する.md'),
    ('ヘルプページ/ヘルプページ_作業用（チェックリスト付き）.md', '20260803_ヘルプページ_CSV一括登録_ドラフト.md'),
    ('仕様書/仕様書（PdM向け）.md', '20260803_CSV一括登録機能_仕様書（PdM向け）.md'),
    ('仕様書/引き継ぎノート.md', 'HANDOFF.md'),
    ('仕様書/判断の記録.md', 'JUDGMENT_LOG.md'),
]
if os.path.isdir(DESK):
    for desk, repo in PAIRS:
        a, b = read(desk, DESK), read(repo)
        if not a:
            ng('配布物', f'デスクトップに {desk} が無い')
        elif a != b:
            ng('配布物', f'{desk} がリポジトリの {repo} と違う')

    # エクセルは元になる md/tsv より新しいこと（古ければ作り直しが必要）
    DERIVED = [
        ('ヘルプページ/ヘルプページ.xlsx', ['20260804_Notion掲載用_CSVで一括登録する.md',
                                            '20260803_ヘルプページ_CSV一括登録_ドラフト.md']),
        ('20260805_タイムレコード_01_受入条件表.xlsx', ['20260804_受け入れ条件_確認表.tsv',
                                                        '20260804_受け入れ条件_定義表.tsv']),
        ('20260805_タイムレコード_02_検証プラン.xlsx', ['20260803_検証プラン_操作系.tsv',
                                                        '20260803_検証プラン_計算系.tsv']),
        ('20260805_タイムレコード_03_検証用データセット.xlsx', ['20260803_検証用データセット.tsv']),
    ]
    for x, srcs in DERIVED:
        px = os.path.join(DESK, x)
        if not os.path.exists(px):
            ng('配布物', f'デスクトップに {x} が無い')
            continue
        for s in srcs:
            ps = os.path.join(B, s)
            if os.path.exists(ps) and os.path.getmtime(ps) > os.path.getmtime(px):
                ng('配布物', f'{x} が {s} より古い（作り直しが必要）')
else:
    print(f'（デスクトップ {DESK} が見つからないため配布物の確認は省略）\n')



# ---------------------------------------------------------------- 8. 検証プランの中身
def plan_rows():
    return [(ho, r) for r in OP] + [(ha, r) for r in CA]


def pv(header, r, name):
    i = col(header, name)
    return r[i] if i is not None else ''


SCREEN_KEY = {'確認画面': '確認画面', '完了画面': '完了画面', 'アップロード画面': 'アップロード',
              '従業員一覧': '従業員一覧', '管理者一覧': '管理者一覧', '業態設定': '業態設定',
              '店舗設定': '店舗設定', '雇用区分設定': '雇用区分設定', 'インポートTOP': 'インポート',
              '初期セットアップ画面': '初期セットアップ', 'ダッシュボード': 'ダッシュボード',
              '実行中の画面': '実行中', '各設定画面': '設定'}
_pr = plan_rows()
_ids = {pv(h, r, '検証ID') for h, r in _pr}

for h, r in _pr:
    v = pv(h, r, '検証ID')
    exp = pv(h, r, '期待挙動（判定基準）')
    scr = pv(h, r, '画面（どこを開いて操作するか）')
    evi = pv(h, r, '証跡（撮る画面と状態）')
    # 空欄が無いこと
    for name in ['検証内容', '初期設定（前提データ）', '期待挙動（判定基準）', 'スクショファイル名',
                 '証跡（撮る画面と状態）', '画面（どこを開いて操作するか）', '確認ポイント',
                 '対応する条件ID', 'ヘルプ該当箇所']:
        if not pv(h, r, name).strip():
            ng('検証プラン', f'{v} の「{name}」が空欄')
    # 期待挙動の【判定する画面】が、画面列か証跡列に出てくること
    m = re.match(r'【([^】]+)】', exp)
    if not m:
        ng('検証プラン', f'{v} の期待挙動に【判定する画面】が無い')
    else:
        for lab in m.group(1).split('・'):
            k = SCREEN_KEY.get(lab)
            if k and k not in scr and k not in evi:
                ng('検証プラン', f'{v} 期待挙動の【{lab}】が画面列にも証跡列にも出てこない')
    # 先に実施する検証
    for d_ in re.findall(r'(?:OP|CA)-\d+', pv(h, r, '先に実施する検証')):
        if d_ not in _ids:
            ng('検証プラン', f'{v} が存在しない {d_} を前提にしている')
        if d_ == v:
            ng('検証プラン', f'{v} が自分自身を前提にしている')

# スクショファイル名
_names = collections.Counter(pv(h, r, 'スクショファイル名') for h, r in _pr)
for nm, cnt in _names.items():
    if cnt > 1:
        ng('検証プラン', f'スクショファイル名が重複: {nm}（{cnt}件）')
    if nm and not nm.endswith('.png'):
        ng('検証プラン', f'スクショファイル名の拡張子が.pngでない: {nm}')
    if re.search(r'[\\/:*?"<>|]', nm):
        ng('検証プラン', f'スクショファイル名に使えない文字: {nm}')

# ---------------------------------------------------------------- 9. 紐付けの双方向一致
plan2ac = {pv(h, r, '検証ID'): set(re.findall(r'AC-\d+', pv(h, r, '対応する条件ID'))) for h, r in _pr}
ac2plan = collections.defaultdict(set)
for r in C:
    for m in re.findall(r'(?:OP|CA)-\d+', r[ievi]):
        ac2plan[m].add(r[0])
for v, s in plan2ac.items():
    if s != ac2plan.get(v, set()):
        ng('紐付け', f'{v} 条件IDが片側だけ プラン={sorted(s)} 受入条件={sorted(ac2plan.get(v, set()))}')

ihelp_c = col(hc, 'ヘルプ該当箇所')
_help_of = {pv(h, r, '検証ID'): pv(h, r, 'ヘルプ該当箇所') for h, r in _pr}
if ihelp_c is not None:
    for r in C:
        exp = set()
        for m in re.findall(r'(?:OP|CA)-\d+', r[ievi]):
            for part in _help_of.get(m, '').split('／'):
                p_ = part.strip()
                if p_ and not p_.startswith('—'):
                    exp.add(p_)
        got = {x.strip() for x in r[ihelp_c].split('／') if x.strip() and not x.strip().startswith('—')}
        if exp != got:
            ng('紐付け', f'{r[0]} ヘルプ該当箇所が検証プランと違う 受入条件={sorted(got)} プラン由来={sorted(exp)}')

# データセットの双方向
# 検証プランは「D1」のようにまとめて参照することがある。その場合 D1-* 全体を指す
_d2v = collections.defaultdict(set)
_grp2v = collections.defaultdict(set)
for h, r in _pr:
    init = pv(h, r, '初期設定（前提データ）')
    v_ = pv(h, r, '検証ID')
    for m in re.findall(r'D\d+-\d+', init):
        _d2v[m].add(v_)
    for gmatch in re.findall(r'D([0-9])(?!-)', init):      # D1〜D5 のまとめ参照
        _grp2v['D' + gmatch].add(v_)
for r in DS:
    did = r[ids_]
    grp = did.split('-')[0]
    covered = _d2v.get(did, set()) | _grp2v.get(grp, set())
    side = set(re.findall(r'(?:OP|CA)-\d+', r[iuse]))
    if _d2v.get(did, set()) - side:
        ng('紐付け', f'{did} の用途欄に {sorted(_d2v[did] - side)} が書かれていない')
    if side - covered:
        ng('紐付け', f'{did} を使うと書かれた {sorted(side - covered)} が検証プランに書かれていない')

# ---------------------------------------------------------------- 10. エクセル ↔ TSV（1行ずつ）
def xlsx_sheet(name):
    p = os.path.join(DESK, name)
    if not os.path.exists(p):
        return None
    try:
        from openpyxl import load_workbook
    except ImportError:
        return None
    return load_workbook(p).active


if os.path.isdir(DESK):
    ws = xlsx_sheet('20260805_タイムレコード_02_検証プラン.xlsx')
    if ws is not None:
        hx = {str(ws.cell(2, i).value): i for i in range(1, ws.max_column + 1)}
        tmap = {pv(h, r, '検証ID'): (h, r) for h, r in _pr}
        if ws.max_row - 2 != len(_pr):
            ng('エクセル', f'02_検証プランの件数 {ws.max_row - 2} ≠ TSV {len(_pr)}')
        cols = ['検証内容', '期待挙動（判定基準）', '画面（どこを開いて操作するか）', 'スクショファイル名',
                '証跡（撮る画面と状態）', '対応する条件ID', 'ヘルプ該当箇所', '先に実施する検証']
        for rr in range(3, ws.max_row + 1):
            k = ws.cell(rr, hx['検証ID']).value
            if k not in tmap:
                ng('エクセル', f'02_検証プランにTSVに無い {k}')
                continue
            h, r = tmap[k]
            for cname in cols:
                if cname not in hx:
                    ng('検査の不備', f'02_検証プランに列「{cname}」が無い')
                    continue
                x = str(ws.cell(rr, hx[cname]).value or '').strip()
                y = pv(h, r, cname).replace('**', '').strip()
                if x != y:
                    ng('エクセル', f'02_検証プラン {k} の「{cname}」がTSVと違う')

    ws = xlsx_sheet('20260805_タイムレコード_01_受入条件表.xlsx')
    if ws is not None:
        hx = {str(ws.cell(28, i).value): i for i in range(1, ws.max_column + 1)}
        dmap = {r[1]: r for r in D}
        PAIR = [('実装レベル', lambda k: c[k][2]),
                ('何を実装するか', lambda k: c[k][4]),
                ('合格ライン（この数値を満たせば実装完了）', lambda k: c[k][6]),
                ('実装できていないと起きること', lambda k: dmap[k][6]),
                ('フロー', lambda k: c[k][1]),
                ('確認方法', lambda k: c[k][7]),
                ('対応検証ID', lambda k: c[k][8]),
                ('ヘルプ該当箇所', lambda k: c[k][ihelp_c] if ihelp_c is not None else '')]
        found = 0
        for rr in range(29, ws.max_row + 1):
            k = ws.cell(rr, hx.get('条件ID', 4)).value
            if k not in c:
                continue
            found += 1
            for cname, fn in PAIR:
                if cname not in hx:
                    ng('検査の不備', f'01_受入条件表に列「{cname}」が無い')
                    continue
                x = str(ws.cell(rr, hx[cname]).value or '').strip()
                y = str(fn(k)).replace('**', '').strip()
                if x != y:
                    ng('エクセル', f'01_受入条件表 {k} の「{cname}」がTSVと違う')
        if found != len(C):
            ng('エクセル', f'01_受入条件表の件数 {found} ≠ TSV {len(C)}')

    ws = xlsx_sheet('20260805_タイムレコード_03_検証用データセット.xlsx')
    if ws is not None:
        hx = {str(ws.cell(1, i).value): i for i in range(1, ws.max_column + 1)}
        start = 2
        if 'データセットID' not in hx:
            hx = {str(ws.cell(2, i).value): i for i in range(1, ws.max_column + 1)}
            start = 3
        dm = {r[ids_]: r for r in DS}
        for rr in range(start, ws.max_row + 1):
            k = ws.cell(rr, 1).value
            if k not in dm:
                continue
            for cname in ['CSVに入力する値', '用途・対応する検証ID']:
                if cname not in hx:
                    ng('検査の不備', f'03_検証用データセットに列「{cname}」が無い')
                    continue
                x = str(ws.cell(rr, hx[cname]).value or '').strip()
                y = dm[k][col(hds, cname)].replace('**', '').strip()
                if x != y:
                    ng('エクセル', f'03_検証用データセット {k} の「{cname}」がTSVと違う')

# マークダウンの記号がエクセルに残っていないか
if os.path.isdir(DESK):
    for nm in ['20260805_タイムレコード_01_受入条件表.xlsx', '20260805_タイムレコード_02_検証プラン.xlsx']:
        cells = xlsx_cells(nm)
        if cells:
            bad = [v for _, _, _, v in cells if isinstance(v, str) and '**' in v]
            if bad:
                ng('エクセル', f'{nm} に ** が {len(bad)}セル残っている（太字に変換すること）')

# ---------------------------------------------------------------- 出力
print('=' * 72)
print('  実データ')
print('=' * 72)
for k, v in N.items():
    print(f'  {k:<16} {v}')
print()
print('=' * 72)
print('  検査結果')
print('=' * 72)
if not NG:
    print('  整合しています。')
    sys.exit(0)
g = collections.defaultdict(list)
for cat, msg in NG:
    g[cat].append(msg)
for cat, msgs in g.items():
    print(f'\n  ■ {cat}（{len(msgs)}件）')
    for m in msgs:
        print(f'    - {m}')
print(f'\n  合計 {len(NG)} 件')
sys.exit(1)
